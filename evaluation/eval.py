#!/usr/bin/env python3
"""Simple evaluation runner for MemOS-lite and RAG baseline.

What it does:
- reads golden QA cases from Excel/CSV;
- runs retrieval + LLM generation WITHOUT QA cache;
- uses touch=False so evaluation does not update access_count;
- computes answer metrics: Token F1, ROUGE-L, answerability accuracy;
- computes retrieval metrics: Hit@1/3/5, MRR, context precision@5, evidence overlap;
- exports compact tables for review.

Expected Excel sheet columns:
    id, question, category, answerable, reference_answer,
    gold_source, gold_ref, gold_quote, note

Optional columns supported:
    gold_memory_id, enabled, review_status

Run:
    python evaluation/eval.py \
      --dataset evaluation/datasets/golden_qa.xlsx \
      --home ./.memos_lite_data \
      --top-k 5 \
      --snapshot-data
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
import time
import unicodedata
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

try:
    from openpyxl import Workbook, load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit("Missing dependency: openpyxl. Install with: pip install openpyxl") from exc

# Allow running from evaluation/eval.py without installing the package.
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.client_factory import get_llm_client  # noqa: E402
from src.config import Config  # noqa: E402
from src.memory_store import MemoryStore  # noqa: E402
from src.retriever import MemoryRetriever  # noqa: E402
from src.schemas import MemoryUnit, Provenance, RetrievedMemory, SourceType  # noqa: E402
from baseline.baseline_qa import BaselineQAService  # noqa: E402
from baseline.baseline_store import BaselineVectorStore  # noqa: E402


PROMPT_TEMPLATE = """Bạn là trợ lý hỏi đáp nghiệp vụ.

Yêu cầu:
- Chỉ trả lời dựa trên NGỮ CẢNH được cung cấp.
- Chỉ nói chưa tìm thấy thông tin khi không có đoạn tài liệu nào liên quan trực tiếp tới câu hỏi.
- Không bịa quy trình, phí, điều kiện, thời hạn, hoặc chính sách.
- Trả lời ngắn gọn, đúng trọng tâm câu hỏi cần trả lời, trả lời bằng tiếng Việt.
- Trong ngữ cảnh có thể có các mục FAQ dạng "Q: ... / A: ...".
  + "Q:" là câu hỏi mẫu trong tài liệu, dùng để nhận biết mục FAQ đó có liên quan với câu hỏi người dùng hay không.
  + Không trả lời lại câu "Q:" trong FAQ như thể đó là câu hỏi chính.
  + Nếu câu hỏi người dùng có ý nghĩa tương đương hoặc gần với "Q:" trong FAQ, hãy ưu tiên dùng nội dung "A:" để trả lời.


### Ngữ cảnh:
{context}

### Câu hỏi cần trả lời:
{question}

### Trả lời:"""

NO_CONTEXT_ANSWER = "Chưa có thông tin trong kho tri thức để trả lời câu hỏi này."
TOKEN_RE = re.compile(r"[\wÀ-ỹ]+", re.UNICODE)
NO_INFO_RE = re.compile(
    r"(chưa\s+(có|tìm\s+thấy)|không\s+(có|tìm\s+thấy|đủ)\s+thông\s+tin|"
    r"không\s+thể\s+xác\s+định|tài\s+liệu\s+không\s+(nêu|đề\s+cập))",
    re.IGNORECASE,
)


@dataclass
class EvalCase:
    case_id: str
    question: str
    category: str = ""
    answerable: bool = True
    reference_answer: str = ""
    gold_source: str = ""
    gold_ref: str = ""
    gold_quote: str = ""
    gold_memory_id: str = ""
    note: str = ""


# ---------------------------------------------------------------------------
# Text normalization and metrics
# ---------------------------------------------------------------------------


def strip_accents(text: str) -> str:
    text = unicodedata.normalize("NFD", text or "")
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn")


def norm_text(text: str) -> str:
    text = strip_accents(text).lower()
    text = re.sub(r"[^\w\s]+", " ", text, flags=re.UNICODE)
    return re.sub(r"\s+", " ", text).strip()


def tokens(text: str) -> list[str]:
    return TOKEN_RE.findall(norm_text(text))


def token_f1(prediction: str, reference: str) -> float:
    pred_toks = tokens(prediction)
    ref_toks = tokens(reference)
    if not pred_toks and not ref_toks:
        return 1.0
    if not pred_toks or not ref_toks:
        return 0.0
    common = Counter(pred_toks) & Counter(ref_toks)
    num_same = sum(common.values())
    if num_same == 0:
        return 0.0
    precision = num_same / len(pred_toks)
    recall = num_same / len(ref_toks)
    return 2 * precision * recall / (precision + recall)


def rouge_l(prediction: str, reference: str) -> float:
    pred_toks = tokens(prediction)
    ref_toks = tokens(reference)
    if not pred_toks and not ref_toks:
        return 1.0
    if not pred_toks or not ref_toks:
        return 0.0

    # LCS dynamic programming, memory-efficient enough for short QA answers.
    prev = [0] * (len(ref_toks) + 1)
    for p in pred_toks:
        curr = [0]
        for j, r in enumerate(ref_toks, start=1):
            curr.append(prev[j - 1] + 1 if p == r else max(prev[j], curr[-1]))
        prev = curr
    lcs = prev[-1]
    precision = lcs / len(pred_toks)
    recall = lcs / len(ref_toks)
    if precision + recall == 0:
        return 0.0
    return 2 * precision * recall / (precision + recall)


def token_recall(needle: str, haystack: str) -> float:
    """How much of `needle` is covered by `haystack`."""

    need = tokens(needle)
    have = tokens(haystack)
    if not need:
        return 0.0
    common = Counter(need) & Counter(have)
    return sum(common.values()) / len(need)


def is_no_info_answer(answer: str) -> bool:
    return bool(NO_INFO_RE.search(answer or ""))


def parse_bool(value: Any, default: bool = True) -> bool:
    if value is None or str(value).strip() == "":
        return default
    text = norm_text(str(value))
    if text in {"yes", "y", "true", "1", "co", "answerable", "duoc"}:
        return True
    if text in {"no", "n", "false", "0", "khong", "unanswerable", "khong duoc"}:
        return False
    return default


# ---------------------------------------------------------------------------
# Dataset IO
# ---------------------------------------------------------------------------


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def canonical_header(value: Any) -> str:
    return norm_text(clean_cell(value)).replace(" ", "_")


def load_cases(path: Path, *, limit: int = 0, skip_unapproved: bool = False) -> list[EvalCase]:
    rows = read_table(path)
    cases: list[EvalCase] = []

    for idx, row in enumerate(rows, start=1):
        enabled = parse_bool(row.get("enabled", "yes"), default=True)
        if not enabled:
            continue
        if skip_unapproved and row.get("review_status") and norm_text(row.get("review_status")) != "approved":
            continue

        question = clean_cell(row.get("question"))
        if not question:
            continue

        case_id = clean_cell(row.get("id")) or f"case_{idx:04d}"
        cases.append(
            EvalCase(
                case_id=case_id,
                question=question,
                category=clean_cell(row.get("category")),
                answerable=parse_bool(row.get("answerable"), default=True),
                reference_answer=clean_cell(row.get("reference_answer")),
                gold_source=clean_cell(row.get("gold_source")),
                gold_ref=clean_cell(row.get("gold_ref")),
                gold_quote=clean_cell(row.get("gold_quote")),
                gold_memory_id=clean_cell(row.get("gold_memory_id")),
                note=clean_cell(row.get("note")),
            )
        )
        if limit and len(cases) >= limit:
            break

    if not cases:
        raise ValueError(f"No valid cases found in {path}")
    return cases


def read_table(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["cases"] if "cases" in wb.sheetnames else wb[wb.sheetnames[0]]
        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows:
            return []
        headers = [canonical_header(v) for v in raw_rows[0]]
        rows: list[dict[str, str]] = []
        for raw in raw_rows[1:]:
            row = {headers[i]: clean_cell(raw[i]) if i < len(raw) else "" for i in range(len(headers)) if headers[i]}
            rows.append(row)
        return rows

    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            return [{canonical_header(k): clean_cell(v) for k, v in row.items()} for row in reader]

    raise ValueError("Dataset must be .xlsx, .xlsm, or .csv")


# ---------------------------------------------------------------------------
# Retrieval matching
# ---------------------------------------------------------------------------


def split_multi(value: str) -> list[str]:
    return [x.strip() for x in str(value or "").split("||") if x.strip()]


def memory_source(item: RetrievedMemory) -> str:
    return item.memory.source or ""


def memory_ref(item: RetrievedMemory) -> str:
    prov = item.memory.provenance
    return prov.source_ref if prov else ""


def source_type(item: RetrievedMemory) -> str:
    prov = item.memory.provenance
    return prov.source_type.value if prov else ""


def source_matches(expected: str, actual: str) -> bool:
    if not expected or not actual:
        return False
    left = norm_text(expected)
    right = norm_text(actual)
    return left in right or right in left


def ref_matches(expected: str, actual: str) -> bool:
    if not expected or not actual:
        return False
    left = norm_text(expected)
    right = norm_text(actual)
    return left in right or right in left


def evaluate_retrieved_chunk(case: EvalCase, item: RetrievedMemory, threshold: float) -> tuple[bool, str, float]:
    mem = item.memory
    methods: list[str] = []

    if case.gold_memory_id and mem.id in split_multi(case.gold_memory_id):
        methods.append("memory_id")

    src_ok = any(source_matches(src, memory_source(item)) for src in split_multi(case.gold_source))
    ref_ok = any(ref_matches(ref, memory_ref(item)) for ref in split_multi(case.gold_ref))
    if src_ok:
        methods.append("source")
        if ref_ok:
            methods.append("source_ref")

    quote_overlaps = [token_recall(q, mem.content) for q in split_multi(case.gold_quote)]
    best_overlap = max(quote_overlaps, default=0.0)
    if best_overlap >= threshold:
        methods.append("text_overlap")

    # Strict enough to avoid marking same-file noise as gold, but flexible when
    # source_ref is unavailable and quote overlap is strong.
    is_gold = bool(
        "memory_id" in methods
        or "source_ref" in methods
        or "text_overlap" in methods
        or (src_ok and best_overlap >= max(0.25, threshold * 0.7))
    )
    return is_gold, "+".join(methods), best_overlap


def retrieval_case_metrics(rows: list[dict[str, Any]], case: EvalCase) -> dict[str, Optional[float]]:
    has_gold = bool(case.answerable and (case.gold_quote or case.gold_source or case.gold_ref or case.gold_memory_id))
    if not has_gold:
        return {
            "hit_at_1": None,
            "hit_at_3": None,
            "hit_at_5": None,
            "mrr": None,
            "context_precision_at_5": None,
            "best_evidence_overlap": None,
        }

    gold_ranks = [int(r["rank"]) for r in rows if r.get("is_gold")]
    best_overlap = max((float(r.get("evidence_overlap") or 0.0) for r in rows), default=0.0)
    top5 = [r for r in rows if int(r["rank"]) <= 5]
    denom = max(1, min(5, len(top5)))

    return {
        "hit_at_1": float(any(r <= 1 for r in gold_ranks)),
        "hit_at_3": float(any(r <= 3 for r in gold_ranks)),
        "hit_at_5": float(any(r <= 5 for r in gold_ranks)),
        "mrr": 1.0 / min(gold_ranks) if gold_ranks else 0.0,
        "context_precision_at_5": sum(1 for r in top5 if r.get("is_gold")) / denom,
        "best_evidence_overlap": best_overlap,
    }


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, rows: list[dict[str, Any]], fields: list[str], sheet_name: str = "cases") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(fields)
    for row in rows:
        ws.append([row.get(field, "") for field in fields])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = min(80, max(len(str(cell.value or "")) for cell in col) + 2)
        ws.column_dimensions[letter].width = max(10, max_len)
    wb.save(path)


def mean(values: list[Optional[float]]) -> Optional[float]:
    clean = [float(v) for v in values if v is not None]
    if not clean:
        return None
    return sum(clean) / len(clean)


def round_or_blank(value: Optional[float], ndigits: int = 4) -> Any:
    return "" if value is None else round(float(value), ndigits)


# ---------------------------------------------------------------------------
# Main evaluation
# ---------------------------------------------------------------------------


def make_config(home: Path, args: argparse.Namespace) -> Config:
    config = Config(home_dir=home)
    if args.model:
        config.ollama_model = args.model
    if args.embedding_model:
        config.ollama_embedding_model = args.embedding_model
    if args.max_context_chars:
        config.max_context_chars = args.max_context_chars
    config.top_k = args.top_k

    # Eval never uses QA cache. We do not call QACache at all, but keep these
    # false in config.json to make the run auditable.
    config.qa_cache_enabled = False
    config.qa_semantic_cache_enabled = False
    return config


def setup_run_dir(base: Path) -> Path:
    run_id = datetime.now().strftime("run_%Y%m%d_%H%M%S")
    run_dir = base / run_id
    run_dir.mkdir(parents=True, exist_ok=False)
    return run_dir


def snapshot_home(original_home: Path, run_dir: Path) -> Path:
    target = run_dir / "data_snapshot"
    ignore = shutil.ignore_patterns("*.lock")
    shutil.copytree(original_home, target, ignore=ignore)
    return target


def build_context(retriever: MemoryRetriever, retrieved: list[RetrievedMemory], max_chars: int) -> str:
    if not retrieved:
        return ""
    return retriever.build_context(retrieved, max_chars=max_chars)


def baseline_as_retrieved(hits: list[dict[str, Any]]) -> list[RetrievedMemory]:
    """Convert plain baseline hits so existing retrieval metrics can be reused."""

    retrieved: list[RetrievedMemory] = []
    for rank, hit in enumerate(hits, start=1):
        metadata = hit.get("metadata") or {}
        file_type = str(metadata.get("file_type") or "").lower()
        source_type = SourceType.FAQ if file_type in {"xlsx", "xlsm"} else SourceType.DOC
        source = str(metadata.get("source") or "baseline")
        source_ref = str(metadata.get("source_ref") or "")
        unit = MemoryUnit(
            id=str(hit.get("id") or f"baseline_{rank}"),
            content=str(hit.get("document") or ""),
            source=source,
            category=str(metadata.get("category") or "uncategorized"),
            provenance=Provenance(
                source_type=source_type,
                source_path=str(metadata.get("file_name") or source),
                source_ref=source_ref,
            ),
        )
        retrieved.append(
            RetrievedMemory(
                memory=unit,
                score=float(hit.get("score") or 0.0),
                rank=rank,
            )
        )
    return retrieved


def should_review(qa_row: dict[str, Any]) -> bool:
    if qa_row.get("error"):
        return True
    if qa_row.get("answerable_gold") != qa_row.get("answerable_pred"):
        return True

    if qa_row.get("answerable_gold") is True:
        if float(qa_row.get("hit_at_5") or 0.0) == 0.0:
            return True
        if float(qa_row.get("token_f1") or 0.0) < 0.25:
            return True
        if float(qa_row.get("rouge_l") or 0.0) < 0.25:
            return True

    return False


def evaluate(args: argparse.Namespace) -> Path:
    dataset_path = Path(args.dataset)
    original_home = Path(args.home)
    run_dir = setup_run_dir(Path(args.out_dir))
    home = snapshot_home(original_home, run_dir) if args.snapshot_data else original_home

    config = make_config(home, args)
    llm_client = get_llm_client(config)
    systems = ["memos", "rag"] if args.mode == "both" else [args.mode]

    retriever: Optional[MemoryRetriever] = None
    if "memos" in systems:
        store = MemoryStore(config, embed_fn=llm_client.embed)
        retriever = MemoryRetriever(store, config, embed_fn=llm_client.embed)

    baseline_service: Optional[BaselineQAService] = None
    if "rag" in systems:
        baseline_service = BaselineQAService(
            store=BaselineVectorStore(
                config.baseline_chroma_path,
                collection_name=config.baseline_collection,
            ),
            embed_fn=llm_client.embed,
            generate_fn=lambda prompt: llm_client.generate(
                prompt, temperature=0.0, seed=args.seed
            ),
            top_k=args.top_k,
            max_context_chars=config.max_context_chars,
        )

    cases = load_cases(dataset_path, limit=args.limit, skip_unapproved=args.skip_unapproved)
    qa_rows: list[dict[str, Any]] = []
    retrieval_rows: list[dict[str, Any]] = []
    review_rows: list[dict[str, Any]] = []
    judge_rows: list[dict[str, Any]] = []
    total = len(cases) * len(systems)

    for index, (case, system) in enumerate(
        ((case, system) for case in cases for system in systems),
        start=1,
    ):
        print(f"[{index}/{total}] [{system}] {case.case_id}: {case.question}")
        start = time.perf_counter()
        answer = ""
        context = ""
        retrieved: list[RetrievedMemory] = []
        case_retrieval_rows: list[dict[str, Any]] = []
        error = ""

        try:
            if system == "memos":
                assert retriever is not None
                retrieved = retriever.retrieve(
                    case.question,
                    top_k=args.top_k,
                    category=case.category or None,
                    touch=False,
                    category_is_strict=args.strict_category,
                )
                context = build_context(retriever, retrieved, config.max_context_chars)
                if context:
                    prompt = PROMPT_TEMPLATE.format(context=context, question=case.question)
                    answer = llm_client.generate(prompt, temperature=0.0, seed=args.seed)
                else:
                    answer = NO_CONTEXT_ANSWER
                latency_ms = (time.perf_counter() - start) * 1000
            else:
                assert baseline_service is not None
                result = baseline_service.answer(case.question, top_k=args.top_k)
                answer = result.answer
                context = result.raw_context
                latency_ms = result.latency_ms

                # Extra deterministic query for retrieval metrics only.
                # It is not included in result.latency_ms.
                hits = baseline_service.store.query(
                    baseline_service.embed_fn(case.question),
                    top_k=args.top_k,
                )
                retrieved = baseline_as_retrieved(hits)

            for rank, item in enumerate(retrieved, start=1):
                is_gold, match_method, evidence_overlap = evaluate_retrieved_chunk(
                    case, item, threshold=args.evidence_threshold
                )
                row = {
                    "run_id": run_dir.name,
                    "system": system,
                    "case_id": case.case_id,
                    "rank": rank,
                    "memory_id": item.memory.id,
                    "version": item.memory.version if system == "memos" else "",
                    "score": round(float(item.score), 6),
                    "source_type": source_type(item),
                    "source": memory_source(item),
                    "source_ref": memory_ref(item),
                    "category": item.memory.category,
                    "tier": item.memory.tier.value if system == "memos" else "",
                    "is_gold": is_gold,
                    "match_method": match_method,
                    "evidence_overlap": round(evidence_overlap, 4),
                    "content": item.memory.content,
                }
                case_retrieval_rows.append(row)
                retrieval_rows.append(row)

        except Exception as exc:  # keep the run alive for remaining cases
            error = repr(exc)
            latency_ms = (time.perf_counter() - start) * 1000

        r_metrics = retrieval_case_metrics(case_retrieval_rows, case)
        qa_row = {
            "run_id": run_dir.name,
            "system": system,
            "case_id": case.case_id,
            "question": case.question,
            "category": case.category,
            "answerable_gold": case.answerable,
            "answerable_pred": not is_no_info_answer(answer),
            "reference_answer": case.reference_answer,
            "answer": answer,
            "token_f1": round(token_f1(answer, case.reference_answer), 4),
            "rouge_l": round(rouge_l(answer, case.reference_answer), 4),
            "hit_at_1": round_or_blank(r_metrics["hit_at_1"]),
            "hit_at_3": round_or_blank(r_metrics["hit_at_3"]),
            "hit_at_5": round_or_blank(r_metrics["hit_at_5"]),
            "mrr": round_or_blank(r_metrics["mrr"]),
            "context_precision_at_5": round_or_blank(r_metrics["context_precision_at_5"]),
            "best_evidence_overlap": round_or_blank(r_metrics["best_evidence_overlap"]),
            "retrieved_count": len(retrieved),
            "latency_ms": round(latency_ms, 2),
            "gold_source": case.gold_source,
            "gold_ref": case.gold_ref,
            "gold_quote": case.gold_quote,
            "note": case.note,
            "error": error,
        }
        qa_rows.append(qa_row)

        judge_rows.append(
            {
                "system": system,
                "case_id": case.case_id,
                "question": case.question,
                "reference_answer": case.reference_answer,
                "gold_quote": case.gold_quote,
                "system_answer": answer,
                "retrieved_context": context,
                "note": case.note,
            }
        )

        if should_review(qa_row):
            top_chunks = "\n\n---\n\n".join(
                f"Rank {r['rank']} | score={r['score']} | source={r['source']} | ref={r['source_ref']}\n{r['content']}"
                for r in case_retrieval_rows[: args.top_k]
            )
            review_rows.append({**qa_row, "top_retrieved_chunks": top_chunks})

    summary = build_summary(run_dir.name, qa_rows)
    write_outputs(run_dir, qa_rows, retrieval_rows, review_rows, judge_rows, summary, config, args)
    return run_dir

def build_summary(run_id: str, qa_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    summaries: list[dict[str, Any]] = []
    for system in dict.fromkeys(row["system"] for row in qa_rows):
        rows = [row for row in qa_rows if row["system"] == system]
        answerable = [
            row for row in rows
            if row["answerable_gold"] is True
            and str(row.get("reference_answer", "")).strip()
        ]
        summaries.append(
            {
                "run_id": run_id,
                "system": system,
                "cases": len(rows),
                "avg_token_f1": round_or_blank(mean([float(r["token_f1"]) for r in answerable])),
                "avg_rouge_l": round_or_blank(mean([float(r["rouge_l"]) for r in answerable])),
                "hit_at_1": round_or_blank(mean([none_if_blank(r["hit_at_1"]) for r in rows])),
                "hit_at_3": round_or_blank(mean([none_if_blank(r["hit_at_3"]) for r in rows])),
                "hit_at_5": round_or_blank(mean([none_if_blank(r["hit_at_5"]) for r in rows])),
                "mrr": round_or_blank(mean([none_if_blank(r["mrr"]) for r in rows])),
                "context_precision_at_5": round_or_blank(
                    mean([none_if_blank(r["context_precision_at_5"]) for r in rows])
                ),
                "best_evidence_overlap": round_or_blank(
                    mean([none_if_blank(r["best_evidence_overlap"]) for r in rows])
                ),
                "answerability_accuracy": round_or_blank(
                    mean([float(r["answerable_gold"] == r["answerable_pred"]) for r in rows])
                ),
                "avg_latency_ms": round_or_blank(mean([float(r["latency_ms"]) for r in rows]), 2),
                "errors": sum(1 for r in rows if r.get("error")),
                "review_cases": sum(1 for r in rows if should_review(r)),
            }
        )
    return summaries

def none_if_blank(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    return float(value)


def write_outputs(
    run_dir: Path,
    qa_rows: list[dict[str, Any]],
    retrieval_rows: list[dict[str, Any]],
    review_rows: list[dict[str, Any]],
    judge_rows: list[dict[str, Any]],
    summary: list[dict[str, Any]],
    config: Config,
    args: argparse.Namespace,
) -> None:
    qa_fields = [
        "run_id",
        "system",
        "case_id",
        "question",
        "category",
        "answerable_gold",
        "answerable_pred",
        "reference_answer",
        "answer",
        "token_f1",
        "rouge_l",
        "hit_at_1",
        "hit_at_3",
        "hit_at_5",
        "mrr",
        "context_precision_at_5",
        "best_evidence_overlap",
        "retrieved_count",
        "latency_ms",
        "gold_source",
        "gold_ref",
        "gold_quote",
        "note",
        "error",
    ]
    retrieval_fields = [
        "run_id",
        "system",
        "case_id",
        "rank",
        "memory_id",
        "version",
        "score",
        "source_type",
        "source",
        "source_ref",
        "category",
        "tier",
        "is_gold",
        "match_method",
        "evidence_overlap",
        "content",
    ]
    summary_fields = [
        "run_id",
        "system",
        "cases",
        "avg_token_f1",
        "avg_rouge_l",
        "hit_at_1",
        "hit_at_3",
        "hit_at_5",
        "mrr",
        "context_precision_at_5",
        "best_evidence_overlap",
        "answerability_accuracy",
        "avg_latency_ms",
        "errors",
        "review_cases",
    ]
    review_fields = qa_fields + ["top_retrieved_chunks"]
    judge_fields = [
        "system",
        "case_id",
        "question",
        "reference_answer",
        "gold_quote",
        "system_answer",
        "retrieved_context",
        "note",
    ]

    write_csv(run_dir / "qa_results.csv", qa_rows, qa_fields)
    write_csv(run_dir / "retrieval_results.csv", retrieval_rows, retrieval_fields)
    write_csv(run_dir / "summary.csv", summary, summary_fields)
    write_xlsx(run_dir / "review_cases.xlsx", review_rows, review_fields, sheet_name="review_cases")
    write_xlsx(run_dir / "judge_input.xlsx", judge_rows, judge_fields, sheet_name="judge_input")

    run_config = {
        "args": vars(args),
        "config": config.as_dict(),
        "note": "QA cache is disabled. Modes: memos, rag, both.",
    }
    with (run_dir / "config.json").open("w", encoding="utf-8") as f:
        json.dump(run_config, f, ensure_ascii=False, indent=2)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Evaluate MemOS, RAG baseline, or both.")
    parser.add_argument("--dataset", required=True, help="Path to golden_qa.xlsx/csv")
    parser.add_argument("--mode", choices=["memos", "rag", "both"], default="both")
    parser.add_argument("--home", default="./.memos_lite_data", help="MemOS-lite data directory")
    parser.add_argument("--out-dir", default="evaluation/runs", help="Output run directory base")
    parser.add_argument("--top-k", type=int, default=5)
    parser.add_argument("--limit", type=int, default=0, help="0 means all cases")
    parser.add_argument("--snapshot-data", action="store_true", help="Copy data dir before eval to avoid mutating original DB")
    parser.add_argument("--strict-category", action="store_true", help="Use category as hard filter instead of soft boost")
    parser.add_argument("--skip-unapproved", action="store_true", help="If review_status exists, only run approved rows")
    parser.add_argument("--evidence-threshold", type=float, default=0.55)
    parser.add_argument("--max-context-chars", type=int, default=0, help="0 keeps Config default")
    parser.add_argument("--model", default="", help="Override Ollama generation model")
    parser.add_argument("--embedding-model", default="", help="Override Ollama embedding model")
    parser.add_argument("--seed", type=int, default=42)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    run_dir = evaluate(args)
    print(f"\nDone. Results written to: {run_dir}")
    print(f"- {run_dir / 'summary.csv'}")
    print(f"- {run_dir / 'qa_results.csv'}")
    print(f"- {run_dir / 'retrieval_results.csv'}")
    print(f"- {run_dir / 'review_cases.xlsx'}")
    print(f"- {run_dir / 'judge_input.xlsx'}")


if __name__ == "__main__":
    main()